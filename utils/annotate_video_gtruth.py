# author: Nikesh Ghimire

import argparse
from collections import deque
import tensorflow as tf
import cv2
import numpy as np
import json
import openpyxl as xl
import os 
import csv
import pandas
import shutil
from random import randrange

def frameOut(path, instName):

    try:

        for i in os.listdir(path):
            if i.endswith('.mp4'):
                vid = path + i
                break

        cap = cv2.VideoCapture(vid)
        ret, frame1 = cap.read()
        prvs = cv2.cvtColor(frame1,cv2.COLOR_BGR2GRAY)
        hsv = np.zeros_like(frame1)
        hsv[...,1] = 255

        global frame_num
        frame = 0
    
    except Exception as e:

        print(e)
        pass

    # out = cv2.VideoWriter('test_vid.mp4',-1,1,(600, 600))

    try:
        while(cap.isOpened()):
            ret, frame2 = cap.read()
            next = cv2.cvtColor(frame2,cv2.COLOR_BGR2GRAY)

            flow = cv2.calcOpticalFlowFarneback(prvs,next, None, 0.5, 3, 15, 3, 5, 1.2, 0)

            mag, ang = cv2.cartToPolar(flow[...,0], flow[...,1])
            hsv[...,0] = ang*180/np.pi/2
            hsv[...,2] = cv2.normalize(mag,None,0,255,cv2.NORM_MINMAX)
            rgb = cv2.cvtColor(hsv,cv2.COLOR_HSV2BGR)

            # dense_flow = cv2.addWeighted(frame, 1,rgb, 2, 0)

            cv2.imshow('frame2',rgb)
            k = cv2.waitKey(30) & 0xff
            if k == 27:
                break
            elif k == ord('s'):
                cv2.imwrite('opticalfb.png',frame2)
                cv2.imwrite('opticalhsv.png',rgb)

            #Check playing or not playing from JSON
            
            s_list = compileInterval(path)
            playing = checkPlay(s_list, frame)

            dir_name = path[path[:-1].rfind('/') + 1:-1]

            if playing:
                cv2.imwrite('LSTM/data/' + instName + '/opFlow/op_' + instName + '_' + str(frame_num) + '_p' + '.jpg', rgb)
                cv2.imwrite('LSTM/data/' + instName + '/raw/raw_' + instName + '_' + str(frame_num) + '_p' + '.jpg', frame2)
            else:
                cv2.imwrite('LSTM/data/' + instName + '/opFlow/op_' + instName + '_' + str(frame_num) + '_np' + '.jpg', rgb)
                cv2.imwrite('LSTM/data/' + instName + '/raw/raw_' + instName + '_' + str(frame_num) + '_np' + '.jpg', frame2)
            
            
            frame_num = frame_num + 1
            frame = frame + 1
            prvs = next

        cap.release()
        cv2.destroyAllWindows()
    
    except Exception as e:
        print(e)
        pass

def compileInterval(path, inst_file):
    # getting rid of file extension
    inst_file = inst_file[:-4]
    for i in os.listdir(path):
        if i.endswith('.JSON') and inst_file in i:
            break
    print(i)
    with open(path + i, 'r', encoding='utf-8-sig') as inFile:
        data = json.load(inFile)

    s_list = [[round(eval(k),2), round(v,2)] for k, v in data['silences'].items()] 
    
    return s_list

def checkPlay(int_list, currFrame):

    currTime = currFrame/30

    for i in int_list:
        if currTime > i[0] and currTime < i[1]:
            return False
        else:
            pass
    
    return True

def compileCSV(path, instName):

    instPath = path + '/' + instName + '/'

    op_list = os.listdir(instPath + '/opFlow/')
    raw_list = os.listdir(instPath + '/raw/')

    wb = xl.Workbook()
    ws = wb.active

    ws.cell(row = 1, column = 1).value = 'name'
    ws.cell(row = 1, column = 2).value = 'label' 

    for i in op_list:

        row_num = ws.max_row  + 1
        ws.cell(row = row_num, column = 1).value = 'opFlow/' + i
        if 'np' in i:
            ws.cell(row = row_num, column = 2).value = '0'
        else:
            ws.cell(row = row_num, column = 2).value = '1'

        print('done with ' + i)

    for i in raw_list:
        row_num = ws.max_row  + 1
        ws.cell(row = row_num, column = 1).value = 'opFlow/' + i
        if 'np' in i:
            ws.cell(row = row_num, column = 2).value = '0'
        else:
            ws.cell(row = row_num, column = 2).value = '1'

        print('done with ' + i)

    wb.save(instPath + '/cleaned_pre.xlsx')

    data_xls = pandas.read_excel(instPath + '/cleaned_pre.xlsx', 'Sheet', index_col=None)
    data_xls.to_csv(instPath + '/cleaned.csv', encoding='utf-8', index=False)

    for i in os.listdir(instPath):
        if i.endswith('.xlsx'):
            os.remove(instPath + '/' + i)

def resize(percent, path, instName):

    instPath = path + instName

    n_playing = os.listdir(instPath + '/n_playing')
    playing = os.listdir(instPath + '/playing')
    test = os.listdir(instPath + '/test')

    moved = 0

    pList = []
    npList = []

    for i in range(round(percent * len(n_playing))):
        
        r_num = randrange(0, len(n_playing))
        x = n_playing[r_num]
        while x in npList:
            r_num = randrange(0, len(n_playing))
            x = n_playing[r_num]
        npList.append(x)
        os.rename(instPath + '/n_playing/' + x, path + instName + '_bckup/bckup/n_playing/' + x)
        moved += 1

    for i in range(round(percent * len(playing))):
        r_num = randrange(0, len(playing))
        x = playing[r_num]
        while x in pList:
            r_num = randrange(0, len(playing))
            x = playing[r_num]
        pList.append(x)
        os.rename(instPath + '/playing/' + x, path + instName + '_bckup/bckup/playing/' + x)
        moved += 1

    print('Moved ' + str(moved) + ' files')

def resize_one(percent, path, instName, pnp):

    instPath = path + instName

    n_playing = os.listdir(instPath + '/n_playing')
    playing = os.listdir(instPath + '/playing')
    test = os.listdir(instPath + '/test')

    moved = 0

    pList = []
    npList = []

    if pnp == 'n':
        for i in range(round(percent * len(n_playing))):
            
            r_num = randrange(0, len(n_playing))
            x = n_playing[r_num]
            while x in npList:
                r_num = randrange(0, len(n_playing))
                x = n_playing[r_num]
            npList.append(x)
            os.rename(instPath + '/n_playing/' + x, path + instName + '_bckup/bckup/n_playing/' + x)
            moved += 1
    elif pnp == 'p':
        for i in range(round(percent * len(playing))):
            r_num = randrange(0, len(playing))
            x = playing[r_num]
            while x in pList:
                r_num = randrange(0, len(playing))
                x = playing[r_num]
            pList.append(x)
            os.rename(instPath + '/playing/' + x, path + instName + '_bckup/bckup/playing/' + x)
            moved += 1

    print('Moved ' + str(moved) + ' files')

def sameSize(path, instName):

    instPath = path + instName

    n_playing = os.listdir(instPath + '/n_playing')
    playing = os.listdir(instPath + '/playing')
    
    moved = 0

    print(len(playing), 'playing')
    print(len(n_playing), 'n_playing')

    pList = []

    for i in range(len(playing) - len(n_playing)):
        
        r_num = randrange(0, len(playing))
        x = playing[r_num]
        while x in pList:
            r_num = randrange(0, len(playing))
            x = playing[r_num]
        pList.append(x)
        os.rename(instPath + '/playing/' + x, path + instName + '_bckup/bckup/playing/' + x)
        moved += 1

    print('Resized playing w.r.t. n_playing --> Moved ' + str(moved) + ' files')

def move(instName, path, n_path, num_files, pnp):

    instPath = path + instName

    n_playing = os.listdir(instPath + '/n_playing')
    playing = os.listdir(instPath + '/playing')
    
    moved = 0

    pList = []

    for i in range(num_files):
        
        r_num = randrange(0, len(playing))
        x = playing[r_num]
        while x in pList:
            r_num = randrange(0, len(playing))
            x = playing[r_num]
        pList.append(x)
        os.rename(instPath + '/playing/' + x, n_path + '/n_playing/cello_moved_' + str(moved) + '.jpg')
        moved += 1

    print('Resized playing w.r.t. n_playing --> Moved ' + str(moved) + ' files')
    

def extract_flow(inst):

    for i in os.listdir('cnn_set/' + inst +'/'):
        n_path = 'cnn_set/' + inst + '/' + i + '/'
        frameOut(n_path, inst)

    # compileCSV('URMP/data/', inst)

def gTruthGene(vidPath, outPath, path):

    vs = cv2.VideoCapture(vidPath)
    # mean = np.array([123.68, 116.779, 103.939][::1], dtype="float32")
    writer = None
    (W, H) = (None, None)
    num = 0

    while True:
        # read the next frame from the file
        (grabbed, frame) = vs.read()
        

        # if the frame was not grabbed, then we have reached the end
        # of the stream
        if not grabbed:
            break

        # if the frame dimensions are empty, grab them
        if W is None or H is None:
            (H, W) = frame.shape[:2]

        # clone the output frame, then convert it from BGR to RGB
        # ordering, resize the frame to a fixed 224x224, and then
        # perform mean subtraction

        output = frame.copy()
        # frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        # frame = cv2.resize(frame, (H + 300, W + 300)).astype("float32")
        # frame -= mean

        # draw the activity on the output frame
        vid_file = vidPath.split('/')[-1]
        s_list = compileInterval(path, vid_file)
        playing = checkPlay(s_list, num)

        print(s_list)
        print(playing)
        num+=1

        if playing:
            text = '---- playing ----'
        else:
            text = '--- n_playing ---'
        
        cv2.putText(output, text, (10, 25), cv2.FONT_HERSHEY_SIMPLEX,
            0.5, (0, 255, 0), 1)


        # check if the video writer is None
        if writer is None:
            # initialize our video writer
            fourcc = cv2.VideoWriter_fourcc(*"MJPG")
            writer = cv2.VideoWriter(outPath, fourcc, 30,
                (W, H), True)

        # write the output frame to disk
        writer.write(output)

        # show the output image
        cv2.imshow("Output", output)
        key = cv2.waitKey(1) & 0xFF

        # if the `q` key was pressed, break from the loop
        if key == ord("q"):
            break

        

    # release the file pointers
    print("[INFO] cleaning up...")
    writer.release()
    vs.release()


def main():

    path = '/home/camel/Documents/URMP_audio_processing/LSTM_dataset_2/violin/20_Pavane_tpt_vn_vc/'
    gTruthGene(path + 'violin_2.mp4', path + 'g_truth_violin_2.avi', path)

    # global frame_num
    # frame_num = 0 

    # extract_flow('violin')

    # frame_num = 0 
    # extract_flow('viola')

    # frame_num = 0 
    # extract_flow('cello')
    # print(compileInterval('cnn_set/trombone/07_GString_tpt_tbn/'))
    # print(compileInterval('cnn_set/trombone/07_GString_tpt_tbn/'))
    

    # frameOut('cnn_set/trombone/07_GString_tpt_tbn/', 'trombone')

    # resize_one(0.15, 'URMP/data/', 'violin_cello_vNN', 'n')
    # sameSize('URMP/data/', 'trumpet')
    # move('cello', 'URMP/data/', 'URMP/data/violin_cello_vNN/', 4385, 'n_playing')

    # compileCSV('LSTM/data/', 'cello')
    # compileCSV('LSTM/data/', 'violin')
    # compileCSV('LSTM/data/', 'viola')
    # compileCSV('LSTM/data/', 'double_bass')compileCSV('LSTM/data/', 'cello')
    # compileCSV('LSTM/data/', 'violin')
    # compileCSV('LSTM/data/', 'viola')
    # compileCSV('LSTM/data/', 'double_bass')

if __name__ == '__main__':

    main()